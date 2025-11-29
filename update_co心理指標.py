import os
import warnings
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
import google.generativeai as genai


# -----------------------------------------
# Gemini（2.5 flash）初期化（あなた指定のコード）
# -----------------------------------------
def init_gemini():
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise RuntimeError("環境変数 GEMINI_API_KEY が設定されていません")
    genai.configure(api_key=api_key)
    return genai.GenerativeModel("gemini-2.5-flash")


# -----------------------------------------
# Schwartz PVQ 定義
# -----------------------------------------
pvq_traits = [
    '自己方向性', '刺激', '享楽', '達成', '権力',
    '安全', '順応', '伝統', '博愛', '普遍主義'
]
pvq_columns = [f'PVQ_{t}' for t in pvq_traits]


# -----------------------------------------
# PVQ推定（Gemini 2.5 flash）
# -----------------------------------------
def extract_pvq_from_value(value_text):
    model = init_gemini()  # 毎回初期化でメモリリーク防止

    prompt = f"""
    あなたは心理学の専門家です。
    以下の文章は、ある企業の「バリュー」または「行動指針」を要約したものです。

    Schwartzの10価値観（PVQ）理論に基づいて、この文章が各価値観をどの程度重視しているかを、
    1〜7 の範囲で推定してください。

    出力形式（順番厳守）：
    自己方向性: 数値
    刺激: 数値
    享楽: 数値
    達成: 数値
    権力: 数値
    安全: 数値
    順応: 数値
    伝統: 数値
    博愛: 数値
    普遍主義: 数値

    ---
    {value_text}
    """

    try:
        res = model.generate_content(prompt)

        if not hasattr(res, "text"):
            return {}

        lines = res.text.strip().splitlines()
        scores = {}

        for line in lines:
            if ":" in line:
                key, val = line.split(":", 1)
                key = key.strip().replace(" ", "")
                if key in pvq_traits:
                    try:
                        scores[f"PVQ_{key}"] = int(val.strip())
                    except:
                        continue

        return scores

    except Exception as e:
        warnings.warn(f"Gemini PVQ推定エラー: {e}")
        return {}


# -----------------------------------------
# Cloud Run Functions 本体  
# worksheet は main() から渡される
# -----------------------------------------
def update_co個人価値観(worksheet):

    print("=== 開始: update_co個人価値観 ===")

    # Colab と同様の方法で DataFrame を作成
    df = worksheet.get_all_records()
    df = pd.DataFrame(df)
    df.fillna("", inplace=True)

    update_count = 0

    for idx, row in df.iterrows():

        company = row.get("会社名", "")
        value_text = row.get("バリュー", "")

        # ------------------------
        # 対象外処理
        # ------------------------
        if company == "対象外" or value_text in ["対象外", "取得失敗"]:

            # すでに対象外ならスキップ
            if all(str(row.get(col, "")).strip() == "対象外" for col in pvq_columns):
                continue

            # 対象外で上書き
            for col in pvq_columns:
                df.at[idx, col] = "対象外"

            update_count += 1
            print(f"⏭️ 対象外: {company}")
            
            # 1行ずつ更新
            for col in pvq_columns:
                col_idx = df.columns.get_loc(col)
                col_letter = get_column_letter(col_idx + 1)
                worksheet.update(
                    f"{col_letter}{idx+2}:{col_letter}{idx+2}",
                    [[df.at[idx, col]]]
                )

            continue

        # ------------------------
        # すでに全て埋まっていればスキップ
        # ------------------------
        if all(str(row.get(col, "")).strip() not in ["", "対象外"] for col in pvq_columns):
            continue

        # ------------------------
        # PVQ推定
        # ------------------------
        scores = extract_pvq_from_value(value_text)

        if scores:
            for col in pvq_columns:
                df.at[idx, col] = scores.get(col, "")
            update_count += 1
            print(f"✅ PVQ成功: {company}")
        else:
            print(f"⚠️ PVQ失敗: {company}")

        # ------------------------
        # Cloud Run の制限回避のため 1行ずつ更新
        # ------------------------
        for col in pvq_columns:
            col_idx = df.columns.get_loc(col)
            col_letter = get_column_letter(col_idx + 1)
            worksheet.update(
                f"{col_letter}{idx+2}:{col_letter}{idx+2}",
                [[df.at[idx, col]]]
            )

    print(f"=== 完了: {update_count}件 更新 ===")

    return {"updated": update_count}
